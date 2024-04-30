import sqlite3
import tkinter as tk
import openpyxl
from tkinter import ttk,messagebox,filedialog
from datetime import datetime
import os


folderpath_order="order/"
folderpath_payment="payment/"
mydb = sqlite3.connect('abx.db')

cursor=mydb.cursor()

cursor.execute('''CREATE TABLE IF NOT EXISTS add_item(
                    id INTEGER PRIMARY KEY,
                    o_num TEXT NOT NULL,
                    t_price TEXT NOT NULL,
                    g_price TEXT NOT NULL,
                    item TEXT NOT NULL,
                    qty TEXT NOT NULL)''')

cursor.execute('''CREATE TABLE IF NOT EXISTS add_payment(
                    id INTEGER PRIMARY KEY,
                    price TEXT NOT NULL,
                    o_number TEXT NOT NULL)''')


def read_order(file_paths):
    for file_path in file_paths:
        
        try:
            delete="DELETE FROM add_item"
            cursor.execute(delete)
            mydb.commit()
            workbook = openpyxl.load_workbook(folderpath_order+file_path)
            
            sheet = workbook.active
            data_rows = iter(sheet.iter_rows(min_row=2, values_only=True))
            for row in data_rows:
                try:
                    gprice=float(row[7])+float(row[9])-float(row[11])-float(row[12])-float(row[13])-float(row[15])
                    query="INSERT INTO `add_item`(`o_num`, `t_price`,`g_price`,`item`,`qty`) VALUES(?,?,?,?,?)"
                    cursor.execute(query,(row[0],row[7],gprice,row[1],row[5],))
                    mydb.commit()
                except Exception as e:
                    print(f"An error occurred: {e}")
            
        except FileNotFoundError:
            print("File not found. Please provide a valid file path.")
        except Exception as e:
            print(f"An error occurred: {e}")
            
def read_payment(file_paths):
    for file_path in file_paths:
        try:
            delete="DELETE FROM add_payment"
            cursor.execute(delete)
            mydb.commit()
            workbook = openpyxl.load_workbook(folderpath_payment+file_path)
            sheet = workbook.active
            data_rows=iter(sheet.iter_rows(min_row=2,values_only=True))
            for row in data_rows:
                try:
                    query="INSERT INTO add_payment(`price`,`o_number`) values(?,?)"
                    cursor.execute(query,(row[13],row[1]))
                    mydb.commit()
                except Exception as e:
                     print(f"An error occurred: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")
    
        
def load_comparation():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    try:
        qry="SELECT o_num,GROUP_CONCAT(item,',') as item, SUM(t_price) as t_price,SUM(g_price) as g_price,SUM(qty) FROM  add_item group by o_num"
        cursor.execute(qry)
        rows=cursor.fetchall()
        sheet["A1"]="Order Number"
        sheet["B1"]="SKU"
        sheet["C1"]="qty"
        sheet["D1"]="total price"
        sheet["E1"]="grand price"
        sheet["F1"]="payment price"
        sheet["G1"]="oustanding"
        for idx,row in enumerate(rows,start=2):
            query=f"SELECT SUM(price) FROM add_payment WHERE o_number='{row[0]}' order by o_number"
            cursor.execute(query)
            paymentprice=cursor.fetchone()
            if paymentprice[0] is not None:
                outstanding=float(row[3])-float(paymentprice[0])
            else:
                outstanding=0
            
            if(outstanding>0):
                sheet.cell(row=idx,column=1).value=row[0]
                sheet.cell(row=idx,column=2).value=row[1]
                sheet.cell(row=idx,column=3).value=row[4]
                sheet.cell(row=idx,column=4).value=row[2]
                sheet.cell(row=idx,column=5).value=row[3]
                sheet.cell(row=idx,column=6).value=paymentprice[0]
                sheet.cell(row=idx,column=7).value=outstanding
            else:
                sheet.cell(row=idx,column=1).value=row[0]
                sheet.cell(row=idx,column=2).value=row[1]
                sheet.cell(row=idx,column=3).value=row[4]
                sheet.cell(row=idx,column=4).value=row[2]
                sheet.cell(row=idx,column=5).value=row[3]
                sheet.cell(row=idx,column=6).value=paymentprice[0]
                sheet.cell(row=idx,column=7).value=outstanding
        current_date = datetime.now().date()
        file_path = f"Comparation of order and payment({current_date}).xlsx"
        workbook.save(file_path)
        messagebox.showinfo("Success", "Comparation is success export to excel")
                
                    
    except Exception as e:
        print(f"An error occurred: {e}")
        
        
            
        
def read_files_in_folder(folder_path):
    try:
        # List all files in the specified folder
        files = os.listdir(folder_path)
        
        # Filter out only files (excluding subdirectories)
        files = [f for f in files if os.path.isfile(os.path.join(folder_path, f))]
        
        # Print the list of files in the folder
        
        for file in files:
            return files
    
    except FileNotFoundError:
        print(f"Folder not found: {folder_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

order=read_files_in_folder(folderpath_order)
payment=read_files_in_folder(folderpath_payment)

read_order(order)
read_payment(payment)

load_comparation()
